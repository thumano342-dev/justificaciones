from flask import Flask, render_template, request, redirect, url_for, flash, session,send_file
import pandas as pd
import mysql.connector
from io import BytesIO
import os 
from werkzeug.utils import secure_filename
from datetime import datetime, timedelta
import shutil
import tempfile
from openpyxl import load_workbook
from datetime import datetime
import re
import smtplib
from email.message import EmailMessage


app = Flask(__name__)
app.secret_key = 'TU_SECRET_KEY'

UPLOAD_FOLDER = "uploads"
app.config["UPLOAD_FOLDER"] = UPLOAD_FOLDER
ALLOWED_EXTENSIONS = {"xlsx", "xls",}


DESTINATARIOS = [
    "talento.humano3@peopleandtrade.com",
]


def extraer_fecha(nombre_archivo):
    try:
        # Buscar un bloque de 8 dígitos (ejemplo: 01102025)
        match = re.search(r"\d{8}", nombre_archivo)
        if match:
            fecha = datetime.strptime(match.group(), "%d%m%Y")
            # 🔹 Devolver fecha en formato 01/10/2025
            return fecha.strftime("%d/%m/%Y")
        return "día desconocido"
    except Exception:
        return "día desconocido"


def enviar_correo_smtp(archivo_path, nombre_archivo):
    msg = EmailMessage()
    msg["Subject"] = f"CONTROL DIARIO DE ASISTENCIA RMS DEL {extraer_fecha(nombre_archivo)}"
    msg["From"] = "diego.avila@peopleandtrade.com"
    msg["To"] = ", ".join(DESTINATARIOS)
    msg.set_content("Buen día, adjunto el control de registro de asistencia en RMS.")
    
    with open(archivo_path, "rb") as f:
        msg.add_attachment(f.read(), maintype="application", subtype="octet-stream", filename=nombre_archivo)

    with smtplib.SMTP("smtp.office365.com", 587) as server:
        server.starttls()
        server.login("diego.avila@peopleandtrade.com", "TU_CONTRASEÑA")
        server.send_message(msg)



# 🔹 Validar extensión de archivo
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


# 📦 CONEXIÓN A MYSQL
def get_connection():
    return mysql.connector.connect(
        host="localhost",
        user="root",
        password="mi2024",
        database="justificaciones"
    )

# 🔐 PINES por canal
PINES = {
    "Carriers": "1234",
    "Entrenamiento": "5678",
    "IR": "2025",
    "Merchandiser": "2024",
    "Open Market": "Hola",
    "Claro": "4567",
    "Admin": "3216",
    "Xiaomi Store": "0000"
    
}

FILES_DIR = r'C:\Attendance\uploads'
ARCHIVO_CONSOLIDADO = os.path.join(UPLOAD_FOLDER, 'Consolidado_Global.xlsx')

# 🔽 OPCIONES SEGÚN RAZÓN
OPCIONES_POR_RAZON = {
    "Sin_Registro_In_Out": [
        '', 'Calamidad Domestica', 'Día Administrativo', 'Día Compensatorio',
        'Día de Familia', 'Día No Remunerado', 'Falla RMS Escalada con soporte Diego Avila',
        'Incapacidad', 'Licencia', 'Sin Justificación',
        'Usuario Retirado', 'Vacaciones', 'Otros'
    ],
    "No_Realizo_Salida": [
        '', 'Sin Justificación', 'Falla RMS Escalada con soporte', 'Otros'
    ],
    "No_cumple_el_Tiempo": [
        '', 'Sin Justificación', 'Otros'
    ]
}

df_global = {}

# 🔹 Validar extensión de archivo
def allowed_file(filename):
    return "." in filename and filename.rsplit(".", 1)[1].lower() in ALLOWED_EXTENSIONS


ADMIN_PASSWORD = "Nacional10" 

@app.route("/admin_login", methods=["GET", "POST"])
def admin_login():
    if request.method == "POST":
        password = request.form.get("password")

        if password == ADMIN_PASSWORD:
            session["is_admin"] = True
            return redirect(url_for("admin_dashboard"))
        else:
            session.pop('_flashes', None)  
            flash("❌ Contraseña incorrecta", "danger")

    return render_template("admin_login.html")

@app.after_request
def add_header(response):
    response.headers["Cache-Control"] = "no-store"
    return response


    

@app.route('/upload', methods=['POST'])
def upload():
    archivo = request.form['archivo']  # o request.files si es un archivo real
    conn = get_connection() 
    cursor = conn.cursor(dictionary=True)

    # 🔹 Verificar si ya existe el archivo en la tabla
    cursor.execute("SELECT * FROM justificaciones WHERE archivo = %s", (archivo,))
    existe = cursor.fetchone()

    if existe:
        # ⚠️ Ya existe → mostrar mensaje
        flash(f"⚠️ El archivo '{archivo}' ya fue subido. Por favor verifica.")
        return redirect('/subir')  # redirige al formulario
    else:
        # 🔹 Insertar porque no existe
        cursor.execute("INSERT INTO justificaciones (archivo) VALUES (%s)", (archivo,))
        conn.commit()
        flash("✅ Archivo subido correctamente")
        return redirect('/subir')

# 📋 DASHBOARD ADMIN
@app.route("/admin_dashboard")
def admin_dashboard():
    if not session.get("is_admin"):
        flash("🔒 Debes iniciar sesión como administrador", "warning")
        return redirect(url_for("admin_login"))
    
    archivos_subidos = [f for f in os.listdir(app.config["UPLOAD_FOLDER"]) if allowed_file(f)]
    archivo_seleccionado = request.args.get("archivo")
    estado_archivo = []

    if archivo_seleccionado:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)
        cursor.execute("""
            SELECT Canal,
                   SUM(CASE WHEN estado = 'pendiente' THEN 1 ELSE 0 END) AS pendientes,
                   SUM(CASE WHEN estado = 'revisado' THEN 1 ELSE 0 END) AS revisados
            FROM justificaciones
            WHERE archivo = %s
            GROUP BY Canal
        """, (archivo_seleccionado,))
        estado_archivo = cursor.fetchall()
        cursor.close()
        conn.close()

    return render_template(
        "admin_dashboard.html",
        archivos=archivos_subidos,
        archivo_seleccionado=archivo_seleccionado,
        estado_archivo=estado_archivo
    )


from openpyxl import load_workbook
import math

def limpiar_valor(valor):
    if valor is None:
        return None
    if isinstance(valor, float) and math.isnan(valor):
        return None
    if isinstance(valor, str) and valor.strip().lower() in ["nan", "none", "null", ""]:
        return None
    return valor



@app.route("/admin", methods=["GET", "POST"])
def admin():
    if not session.get("is_admin"):
        flash("🔒 Debes iniciar sesión como administrador", "warning")
        return redirect(url_for("admin_login"))

    if request.method == "POST":
        if "file" not in request.files:
            flash("No se envió ningún archivo", "danger")
            return redirect(request.url)

        file = request.files["file"]
        if file.filename == "":
            flash("Selecciona un archivo válido", "warning")
            return redirect(request.url)

        if file and allowed_file(file.filename):
            filename = secure_filename(file.filename)

            temp_dir = tempfile.mkdtemp()
            temp_path = os.path.join(temp_dir, filename)
            file.save(temp_path)

            try:
                
                conn = get_connection()
                cursor = conn.cursor()
                cursor.execute("SELECT COUNT(*) FROM justificaciones WHERE archivo = %s", (filename,))
                existe = cursor.fetchone()[0]
                cursor.close()
                conn.close()

                if existe > 0:
                    flash(f"⚠️ El archivo '{filename}' ya fue subido anteriormente. Verifica antes de continuar.", "warning")
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    return redirect(request.url)

                # 🔍 Validar si ya existe físicamente en uploads
                upload_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
                if os.path.exists(upload_path):
                    flash(f"⚠️ El archivo '{filename}' ya existe en el servidor. Verifica antes de continuar.", "warning")
                    shutil.rmtree(temp_dir, ignore_errors=True)
                    return redirect(request.url)

                # ✅ Usamos openpyxl para leer la tabla "TablaJust" en la hoja "Hoja1"
                wb = load_workbook(temp_path, data_only=True)

                if "Hoja1" not in wb.sheetnames:
                    raise ValueError("El archivo no contiene la hoja 'Hoja1'.")

                ws = wb["Hoja1"]

                if "TablaJust" not in ws.tables:
                    raise ValueError("El archivo no contiene la tabla 'TablaJust'.")

                tabla = ws.tables["TablaJust"]
                ref = tabla.ref
                rango = ws[ref]
                data = [[cell.value for cell in row] for row in rango]

                if not data or len(data) < 2:
                    raise ValueError("La tabla 'TablaJust' no tiene datos.")

                columnas = data[0]
                filas = data[1:]
                df = pd.DataFrame(filas, columns=columnas)
                df.columns = df.columns.str.strip()
                df["archivo"] = filename

                columnas_requeridas = ["Canal", "CM", "User Code", "English Name", 
                                       "User Title", "Razon", "Cedula", 
                                       "Justificación", "Observaciones en caso de Otros"]
                for col in columnas_requeridas:
                 if col not in df.columns:
                    raise ValueError(f"❌ La columna requerida '{col}' no existe en el archivo '{filename}'.")


                # 👉 Si todo está bien, mover a uploads
                shutil.move(temp_path, upload_path)

                # ✅ Guardar en MySQL
                conn = get_connection()
                cursor = conn.cursor()
                for _, row in df.iterrows():
                    valores = [limpiar_valor(row.get(col, None)) for col in [
                        "archivo", "Canal", "CM", "User Code", "English Name", "User Title",
                        "Razon", "Cedula", "Justificación", "Observaciones en caso de Otros"
                    ]]
                    placeholders = ", ".join(["%s"] * len(valores))
                    sql = f"""
                        INSERT INTO justificaciones 
                        (archivo, Canal, CM, user_code, english_name, user_title, razon, Cedula, justificacion, observaciones, estado) 
                        VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                    """

                    valores.append("pendiente")  
                    cursor.execute(sql, valores)
                    

                conn.commit()
                cursor.close()
                conn.close()

                # 🧹 Limpiar carpeta temporal
                shutil.rmtree(temp_dir)

                archivo_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)


                

                flash("✅ Archivo cargado, procesado y enviado por correo correctamente.", "success")
                return redirect(url_for("admin"))


            except Exception as e:
                # ❌ Si falla → borrar archivo temporal y carpeta
                if os.path.exists(temp_path):
                    os.remove(temp_path)
                shutil.rmtree(temp_dir, ignore_errors=True)
                flash(f"❌ Error en el archivo: {e}", "danger")
                return redirect(request.url)

    return render_template("admin.html")


                 
                 
                
               


# 🔹 ELIMINAR ARCHIVO
@app.route("/admin/eliminar_archivo", methods=["POST"])
def admin_eliminar_archivo():
    if not session.get("is_admin"):
        flash("🔒 Debes iniciar sesión como administrador", "warning")
        return redirect(url_for("admin_login"))

    archivo = request.form.get("archivo")
    if not archivo:
        flash("❌ No se indicó ningún archivo", "danger")
        return redirect(url_for("admin_dashboard"))

    filepath = os.path.join(app.config["UPLOAD_FOLDER"], archivo)
    if os.path.exists(filepath):
        os.remove(filepath)

    try:
        conn = get_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM justificaciones WHERE archivo = %s", (archivo,))
        conn.commit()
        cursor.close()
        conn.close()
    except Exception as e:
        flash(f"Error al borrar archivo de la base: {e}", "danger")
        return redirect(url_for("admin_dashboard"))

    if archivo in df_global:
        df_global.pop(archivo)

    flash(f"Archivo '{archivo}' eliminado correctamente ✅", "success")
    return redirect(url_for("admin_dashboard"))

# 👋 CERRAR SESIÓN ADMIN
@app.route("/admin_logout")
def admin_logout():
    session.pop("is_admin", None)
    flash("👋 Sesión de administrador cerrada", "info")
    return redirect(url_for("index"))



# 📂 LISTAR ARCHIVOS EN EL DIRECTORIO
def listar_archivos():
    conn = get_connection()
    cursor = conn.cursor()
    cursor.execute("SELECT DISTINCT archivo FROM justificaciones")
    archivos = [row[0] for row in cursor.fetchall()]
    conn.close()
    return archivos


# 📥 CARGAR DATOS DESDE MYSQL
def cargar_datos_mysql(archivo):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Traemos todo de justificaciones y el nombre del canal
    cursor.execute("""
    SELECT j.user_code, j.english_name, j.user_title, 
           j.razon, j.justificacion, j.observaciones, j.estado,
           j.CM, j.archivo,
           j.Canal
    FROM justificaciones j
    WHERE j.archivo = %s
""", (archivo,))
    
    
    rows = cursor.fetchall()
    conn.close()

    df = pd.DataFrame(rows)
    if not df.empty:
        df.rename(columns={
            'user_code': 'User Code',
            'english_name': 'English Name',
            'user_title': 'User Title',
            'razon': 'Razon',
            'justificacion': 'Justificación',
            'observaciones': 'Observaciones en caso de Otros'
        }, inplace=True)
        df = df.fillna('')
        df['Canal'] = df['Canal'].astype(str).str.strip()
    else:
        # Si no hay datos, devolvemos un DF vacío con la columna Canal incluida
        df = pd.DataFrame(columns=['User Code','English Name','User Title','Razon',
                                   'Justificación','Observaciones en caso de Otros',
                                   'CM','archivo','Canal','estado'])
    return df

# 🔹 Ruta para mostrar datos por canal
@app.route("/canal/<canal_name>")
def ver_canal(canal_name):
    try:
        conn = get_connection()
        cursor = conn.cursor(dictionary=True)  # Devuelve filas como diccionarios
        cursor.execute("""
            SELECT archivo, CM, user_code, english_name, user_title, razon, Cedula, 
                   justificacion, observaciones, estado
            FROM justificaciones
            WHERE Canal = %s
        """, (canal_name,))
        registros = cursor.fetchall()
        cursor.close()
        conn.close()

        if not registros:
            flash(f"No hay registros para el canal {canal_name}", "info")

        return render_template("canal.html", canal=canal_name, registros=registros)

    except Exception as e:
        flash(f"Error al obtener datos: {e}", "danger")
        return redirect(url_for("index"))


# 📦 CACHE DE DATAFRAME EN MEMORIA
def cargar_datos_en_memoria(nombre_archivo):
    global df_global
    if nombre_archivo not in df_global:
        df_global[nombre_archivo] = cargar_datos_mysql(nombre_archivo)
    return df_global[nombre_archivo]

# 📤 GUARDAR CAMBIOS EN MYSQL
def actualizar_justificaciones_mysql(archivo, df, edited_rows):
    conn = get_connection()
    cursor = conn.cursor()
    for i in edited_rows:
        just = df.loc[i, 'Justificación']
        obs = df.loc[i, 'Observaciones en caso de Otros']
        user_code = df.loc[i, 'User Code']
        cursor.execute("""
            UPDATE justificaciones 
            SET justificacion = %s, observaciones = %s
            WHERE archivo = %s AND user_code = %s
        """, (just, obs, archivo, user_code))
    conn.commit()
    conn.close()

# 📊 GUARDAR EN EXCEL CONSOLIDADO
def guardar_consolidado(df_nuevo):
    if not os.path.exists(ARCHIVO_CONSOLIDADO):
        df_existente = pd.DataFrame()
    else:
        try:
            df_existente = pd.read_excel(ARCHIVO_CONSOLIDADO)
        except:
            df_existente = pd.DataFrame()
    df_final = pd.concat([df_existente, df_nuevo], ignore_index=True)
    df_final.to_excel(ARCHIVO_CONSOLIDADO, index=False)



@app.route('/descargar_excel/<int:canal_id>')
def descargar_excel(canal_id):
    conn = get_connection()
    cursor = conn.cursor(dictionary=True)

    # Traemos todos los registros de ese canal con el nombre del canal
    cursor.execute("""
        SELECT j.user_code, j.english_name, j.user_title, 
               j.razon, j.justificacion, j.observaciones, 
               j.CM, j.archivo,
               c.nombre AS Canal
        FROM justificaciones j
        LEFT JOIN canales c ON j.canal_id = c.id
        WHERE j.canal_id = %s
    """, (canal_id,))
    
    rows = cursor.fetchall()
    conn.close()

    if not rows:
        flash("No hay datos para exportar en este canal.", "warning")
        return redirect(url_for('index'))

    # Convertir a DataFrame
    df = pd.DataFrame(rows)

    # Renombrar columnas
    df.rename(columns={
        'user_code': 'User Code',
        'english_name': 'English Name',
        'user_title': 'User Title',
        'razon': 'Razon',
        'justificacion': 'Justificación',
        'observaciones': 'Observaciones en caso de Otros'
    }, inplace=True)

    df = df.fillna('')
    df['Canal'] = df['Canal'].astype(str).str.strip()

    # Guardar en memoria como Excel
    from io import BytesIO
    output = BytesIO()
    with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
        df.to_excel(writer, index=False, sheet_name='Datos')

    output.seek(0)

    # Devolver el archivo Excel
    return send_file(
        output,
        as_attachment=True,
        download_name=f'canal_{canal_id}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


@app.route('/salir', methods=['POST'])
def salir():
    session.pop('canal', None)
    flash("Has salido del canal.")
    return redirect(url_for('index'))

# 🌐 RUTA PRINCIPAL
@app.route('/', methods=['GET', 'POST'])
def index():
    canal = request.form.get('canal') or request.args.get('canal') or session.get('canal')
    pin = request.form.get('pin')

    # ✅ VALIDACIÓN DEL PIN
    if request.method == "POST" and 'pin' in request.form and canal:
        if canal == "Admin" and pin == PINES.get("Admin"):
            session['canal'] = canal
            session['canal_autenticado'] = canal
            return redirect(url_for("admin_dashboard"))
        
        elif pin == PINES.get(canal):
            session['canal'] = canal
            session['canal_autenticado'] = canal
            return redirect(url_for("index", canal=canal))
        else:
            flash("❌ PIN incorrecto. Intenta nuevamente.", "danger")
            return render_template("index.html", canales=list(PINES.keys()), canal=canal)

    # ✅ SI SELECCIONA CANAL PERO NO ESTÁ AUTENTICADO
    if canal and session.get("canal_autenticado") != canal:
        flash("🔒 Ingresa el PIN para acceder a este canal.")
        return render_template("index.html", canales=list(PINES.keys()), canal=canal)

    # ✅ SI YA ESTÁ AUTENTICADO → MOSTRAR DATA
    perfil = session.get('canal_autenticado')
    archivos = listar_archivos()
    canal_seleccionado = canal
    archivo_seleccionado = request.args.get('archivo') or request.form.get('archivo')
    filtro_cm = request.args.get('cm')

    progreso_archivos = {}
    for archivo in archivos:
        df = cargar_datos_en_memoria(archivo)
        if 'Canal' in df.columns:
            df_canal = df[df['Canal'] == canal_seleccionado]

            total_rows = len(df_canal)
            df_pendientes = df_canal[df_canal['Justificación'].isna() | (df_canal['Justificación'] == '')]
            completados = total_rows - len(df_pendientes)

            if total_rows == 0:
                progreso = 0
            elif completados == 0:
                progreso = 0
            elif completados == total_rows:
                progreso = 100
            else:
                progreso = round((completados / total_rows) * 100, 1)

            progreso_archivos[archivo] = progreso

    canales = sorted(PINES.keys())

    if not canal_seleccionado:
        return render_template('index.html', canales=canales, progreso_archivos=progreso_archivos)

    archivos_filtrados = [arch for arch in archivos if canal_seleccionado in cargar_datos_en_memoria(arch)['Canal'].values]

    # ✅ Verifica si canal está completamente diligenciado
    canal_completo = True
    for arch in archivos_filtrados:
        df_arch = cargar_datos_en_memoria(arch)
        df_arch['Canal'] = df_arch['Canal'].astype(str).str.strip()
        pendientes = df_arch[(df_arch['Canal'] == canal_seleccionado) &
                             ((df_arch['Justificación'].isna()) | (df_arch['Justificación'] == ''))]
        if not pendientes.empty:
            canal_completo = False
            break

    
        
    df = pd.DataFrame()
    responsables = []
    df_pendientes = pd.DataFrame()
    progreso = completados = total = 0
    edited_rows = []  # 🔹 Siempre existe aunque no se edite nada

    if archivo_seleccionado:
        df = cargar_datos_en_memoria(archivo_seleccionado)
        df['Canal'] = df['Canal'].astype(str).str.strip()
        df = df[df['Canal'] == canal_seleccionado]

        if request.method == 'POST':
            for i in df.index:
                just = request.form.get(f'just_{i}')
                obs = request.form.get(f'obs_{i}')
                if just and just.strip() != "":
                    if just in ['Otros', 'Falla RMS Escalada con soporte Diego Avila', 'Falla RMS Escalada con soporte']:
                        obs = obs if obs and len(obs.strip()) >= 1 else ""
                    df.loc[i, 'Justificación'] = just
                    df.loc[i, 'Observaciones en caso de Otros'] = obs
                    edited_rows.append(i)

    # ✅ Solo si hubo ediciones
    if edited_rows:
        actualizar_justificaciones_mysql(archivo_seleccionado, df, edited_rows)
        guardar_consolidado(df.loc[edited_rows])
        df_global[archivo_seleccionado] = cargar_datos_mysql(archivo_seleccionado)

        # 🔍 Revisar si ya no quedan pendientes
        pendientes = df[df['Justificación'].isna() | (df['Justificación'] == '')]
        if pendientes.empty:
            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("""
                UPDATE justificaciones 
                SET estado = 'revisado'
                WHERE archivo = %s AND Canal = %s
            """, (archivo_seleccionado, canal_seleccionado))
            conn.commit()
            cursor.close()
            conn.close()
            flash("🏆 Canal completado. Se marcó como revisado automáticamente.", "success")

    # 👉 Esto siempre se calcula
    if not df.empty:
        responsables = df['CM'].dropna().unique().tolist()
        if filtro_cm:
            df = df[df['CM'] == filtro_cm]

        df_pendientes = df[df['Justificación'].isna() | (df['Justificación'] == '')]
        total = len(df)
        completados = total - len(df_pendientes)
        progreso = round((completados / total) * 100, 1) if total else 0

    return render_template(
        'index.html',
        canales=canales,
        perfil=canal_seleccionado,
        archivos=archivos_filtrados,
        archivo=archivo_seleccionado,
        responsables=responsables,
        filtro_cm=filtro_cm,
        df_filtrado=list(df_pendientes.iterrows()) if not df_pendientes.empty else [],
        opciones_por_razon=OPCIONES_POR_RAZON,
        progreso=progreso,
        completados=completados,
        total=total,
        progreso_archivos=progreso_archivos,
        
    )

if __name__ == '__main__':
    app.run(debug=True)
